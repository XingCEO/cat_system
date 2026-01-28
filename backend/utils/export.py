"""
Export Service - Export data to various formats
"""
import csv
import json
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to various formats"""
    
    # Default columns for stock export
    DEFAULT_COLUMNS = [
        ("symbol", "股票代號"),
        ("name", "股票名稱"),
        ("industry", "產業類別"),
        ("open_price", "開盤價"),
        ("high_price", "最高價"),
        ("low_price", "最低價"),
        ("close_price", "收盤價"),
        ("change_percent", "漲幅(%)"),
        ("volume", "成交量(張)"),
        ("consecutive_up_days", "連續上漲天數"),
        ("amplitude", "振幅(%)"),
        ("volume_ratio", "量比"),
        ("distance_from_high", "距52週高點(%)"),
        ("distance_from_low", "距52週低點(%)"),
        ("avg_change_5d", "近5日平均漲幅(%)"),
    ]
    
    @classmethod
    def to_csv(
        cls,
        data: List[Dict[str, Any]],
        columns: Optional[List[tuple]] = None
    ) -> str:
        """
        Export data to CSV string
        
        Args:
            data: List of dictionaries to export
            columns: List of (key, display_name) tuples
        """
        if not data:
            return ""
        
        columns = columns or cls.DEFAULT_COLUMNS
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        header = [col[1] for col in columns]
        writer.writerow(header)
        
        # Write data rows
        for row in data:
            values = []
            for key, _ in columns:
                value = row.get(key, "")
                if isinstance(value, float):
                    value = round(value, 2)
                values.append(value if value is not None else "")
            writer.writerow(values)
        
        return output.getvalue()
    
    @classmethod
    def to_excel(
        cls,
        data: List[Dict[str, Any]],
        columns: Optional[List[tuple]] = None,
        sheet_name: str = "篩選結果"
    ) -> bytes:
        """
        Export data to Excel bytes
        
        Args:
            data: List of dictionaries to export
            columns: List of (key, display_name) tuples
            sheet_name: Name of the worksheet
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        if not data:
            return b""
        
        columns = columns or cls.DEFAULT_COLUMNS
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Write header
        for col_idx, (key, display_name) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (key, _) in enumerate(columns, 1):
                value = row_data.get(key, "")
                if isinstance(value, float):
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
                cell.border = thin_border
                
                # Right align numbers
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-adjust column widths
        for col_idx, (key, display_name) in enumerate(columns, 1):
            max_length = len(display_name)
            for row in data:
                value = str(row.get(key, ""))
                max_length = max(max_length, len(value))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @classmethod
    def to_json(
        cls,
        data: List[Dict[str, Any]],
        pretty: bool = True
    ) -> str:
        """
        Export data to JSON string
        
        Args:
            data: List of dictionaries to export
            pretty: Whether to format with indentation
        """
        if pretty:
            return json.dumps(data, ensure_ascii=False, indent=2)
        return json.dumps(data, ensure_ascii=False)
    
    @classmethod
    def generate_filename(
        cls,
        prefix: str = "stocks",
        extension: str = "csv",
        include_timestamp: bool = True
    ) -> str:
        """Generate a filename with optional timestamp"""
        if include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return f"{prefix}_{timestamp}.{extension}"
        return f"{prefix}.{extension}"
    
    @classmethod
    def generate_share_url(
        cls,
        base_url: str,
        params: Dict[str, Any]
    ) -> str:
        """Generate a shareable URL with query parameters"""
        from urllib.parse import urlencode, quote
        
        # Filter out None values
        filtered_params = {k: v for k, v in params.items() if v is not None}
        
        # Handle list parameters
        query_parts = []
        for key, value in filtered_params.items():
            if isinstance(value, list):
                for item in value:
                    query_parts.append(f"{key}={quote(str(item))}")
            else:
                query_parts.append(f"{key}={quote(str(value))}")
        
        query_string = "&".join(query_parts)
        return f"{base_url}?{query_string}"


# Global instance
export_service = ExportService()
